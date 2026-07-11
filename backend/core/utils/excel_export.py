#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
通用 Excel 导出构建器
基于 openpyxl，将 ORM 对象列表导出为 .xlsx 字节流
"""
import io
from dataclasses import dataclass, field
from typing import Any, Callable, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SYNC_EXPORT_MAX_ROWS = 10000


@dataclass
class ExportColumn:
    """导出列定义"""

    field: str
    header: str
    width: int = 20
    table: str | None = None
    transform: Callable[[Any], Any] | None = None
    number_format: str | None = None


def _get_value(row: Any, col: ExportColumn) -> Any:
    """从行中取值，兼容 ORM 对象和字典"""
    if isinstance(row, dict):
        key = f"{col.table}.{col.field}" if col.table else col.field
        val = row.get(key, row.get(col.field))
    else:
        val = getattr(row, col.field, None)
    if col.transform and val is not None:
        val = col.transform(val)
    return "" if val is None else val


def build_excel_bytes(
    columns: Sequence[ExportColumn],
    rows: Sequence[Any],
    sheet_name: str = "Sheet1",
) -> bytes:
    """
    将 ORM 行对象或字典列表构建为 .xlsx 字节。
    支持两种行格式：
    - ORM 对象：通过 getattr(col.field) 取值
    - dict（跨表 JOIN 结果）：通过 dict[col.table.col.field] 或 dict[col.field] 取值
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写表头
    ws.append([col.header for col in columns])

    # 设置列宽
    for idx, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = col.width

    # 写数据行，并应用数值格式
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            value = _get_value(row, col)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col.number_format:
                cell.number_format = col.number_format

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
